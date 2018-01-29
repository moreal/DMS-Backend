from openpyxl import Workbook

from flask import Blueprint, send_from_directory
from flask_jwt_extended import get_jwt_identity, jwt_required
from flask_restful import Api, Resource, abort
from flasgger import swag_from

from app.docs.admin.apply.stay import *
from app.models.account import AdminModel, StudentModel
from utils.apply_excel_manager import get_cells, ready_worksheet

api = Api(Blueprint('admin-stay-api', __name__))
api.prefix = '/admin'


@api.resource('/stay')
class StayDownload(Resource):
    @swag_from(STAY_DOWNLOAD_GET)
    @jwt_required
    def get(self):
        """
        잔류신청 엑셀 다운로드
        """
        admin = AdminModel.objects(id=get_jwt_identity()).first()
        if not admin:
            abort(403)

        wb = Workbook()
        ws = wb.active

        ready_worksheet(ws)

        for student in StudentModel.objects:
            number_cell, name_cell, status_cell = get_cells(student)

            ws[number_cell] = student.number
            ws[name_cell] = student.name

            stay_value = student.stay_apply.value

            if stay_value == 1:
                status = '금요 귀가'
            elif stay_value == 2:
                status = '토요 귀가'
            elif stay_value == 3:
                status = '토요 귀사'
            else:
                status = '잔류'

            ws[status_cell] = status

        wb.save('stay.xlsx')
        wb.close()

        return send_from_directory('../', 'stay.xlsx')
